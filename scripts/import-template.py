#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Импорт ШАБЛОН_personal_budget.xlsx

Лист «Операции»: дата | название | счёт | счёт_куда | сумма
  Категория расхода — по заливке ячейки «название» (см. FILL_TO_CATEGORY).

Лист «Балансы» (необязательно): сверка остатков по месяцам.
Лист «Доходы»: месяц | источник | сумма | тип карты — внешние поступления (переводы пропускаются).
"""

import json
import re
import sqlite3
import sys
import uuid
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

try:
    import openpyxl
    from openpyxl.styles.colors import COLOR_INDEX
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent
DB_PATH = ROOT / "data" / "personal-budget.sqlite"
SCHEMA_PATH = ROOT / "api" / "schema.sql"
RULES_PATH = ROOT / "api" / "rules" / "categorize.json"
SEED_ACCOUNTS = ROOT / "api" / "seed-accounts.json"
SEED_CATEGORIES = ROOT / "api" / "seed-categories.json"
DEFAULT_XLSX = ROOT / "data/template/ШАБЛОН_personal_budget.xlsx"

CREDIT_LIMIT_DEFAULT = 380000.0
# Если в листе «Балансы» стоит полный лимит 380 000 — подставляем фактическое доступно
CREDIT_AVAILABLE_FALLBACK = 348538.0

IMPORT_YEAR = 2026

MONTH_PARTS = [
    ("январ", 1), ("феврал", 2), ("март", 3), ("апрел", 4),
    ("май", 5), ("мая", 5), ("июн", 6), ("июня", 6),
    ("июл", 7), ("июля", 7), ("август", 8), ("сентябр", 9),
    ("октябр", 10), ("ноябр", 11), ("декабр", 12),
]

BALANCE_MONTH_RU = {
    "январ": "01", "феврал": "02", "март": "03", "апрел": "04",
    "май": "05", "мая": "05", "июн": "06", "июня": "06",
    "июл": "07", "июля": "07", "август": "08", "сентябр": "09",
    "октябр": "10", "ноябр": "11", "декабр": "12",
}

CATEGORY_SLUGS = {
    "еда": "food", "зарплата": "salary", "кредиты": "credits", "игры": "games",
    "алкоголь": "alcohol", "ежемесячные": "monthly", "стануша": "stanusha",
    "другое": "other", "marketplace": "marketplace", "авто": "car",
    "здоровье": "health", "дом": "home", "транспорт": "transport",
}

# Заливка ячейки «название» на листе «Операции» (Google Sheets / Excel)
FILL_TO_CATEGORY = {
    "f7cb4d": "alcohol",   # жёлтый — алкоголь
    "f1c232": "alcohol",   # второй жёлтый (редко)
    "d9ead3": "food",      # светло-зелёный — еда
    "cfe2f3": "other",     # светло-синий — остальное
    "d9d2e9": "monthly",   # светло-фиолетовый — ежемесячные
    "f4cccc": "credits",   # светло-красный — кредиты
}


def parse_money(value) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).replace("\xa0", "").replace(" ", "").replace("₽", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def new_id() -> str:
    return str(uuid.uuid4())


def cell_fill_hex(cell) -> Optional[str]:
    fill = cell.fill
    if not fill or fill.fill_type != "solid":
        return None
    color = fill.fgColor
    if color is None:
        return None
    if color.type == "rgb" and color.rgb:
        hx = color.rgb
        if len(hx) == 8:
            hx = hx[2:]
        return hx.lower()
    if color.type == "indexed" and color.indexed is not None:
        try:
            v = COLOR_INDEX[color.indexed]
            if v and len(v) >= 6:
                return v[-6:].lower()
        except (IndexError, KeyError):
            pass
    return None


def category_from_fill(ws, row: int, name_col: int) -> Optional[str]:
    hx = cell_fill_hex(ws.cell(row, name_col))
    if hx:
        return FILL_TO_CATEGORY.get(hx)
    return None


def load_rules():
    if RULES_PATH.is_file():
        return json.loads(RULES_PATH.read_text(encoding="utf-8"))
    return []


def suggest_category(title: str, rules) -> Optional[str]:
    lower = (title or "").lower()
    for rule in rules:
        for kw in rule.get("keywords", []):
            if kw.lower() in lower:
                return rule["categoryId"]
    return None


def normalize_header(value) -> str:
    return re.sub(r"\s+", "_", str(value or "").strip().lower())


def map_account(raw: Optional[str]) -> Optional[str]:
    if not raw:
        return None
    s = str(raw).strip().lower()
    if "кредитк" in s or s == "кредитная карта":
        return "credit_card"
    if "на кредит" in s or "для оплаты кредит" in s:
        return "credit_payments_card"
    if "общ" in s:
        return "shared_card"
    if "основн" in s:
        return "main_card"
    if s in ("основная", "общая", "на_кредиты", "кредитка"):
        return {
            "основная": "main_card",
            "общая": "shared_card",
            "на_кредиты": "credit_payments_card",
            "кредитка": "credit_card",
        }[s]
    return None


def parse_month_cell(value) -> Optional[str]:
    if value is None or str(value).strip() == "":
        return None
    mkey = str(value).strip().lower()
    for part, suffix in BALANCE_MONTH_RU.items():
        if mkey.startswith(part):
            return f"{IMPORT_YEAR}-{suffix}"
    return None


def read_income_headers(ws) -> Dict[str, int]:
    mapping = {}
    for c in range(1, ws.max_column + 1):
        h = normalize_header(ws.cell(1, c).value)
        if h:
            mapping[h] = c
    aliases = {
        "месяц": "месяц",
        "источник_дохода": "источник",
        "источние_дохода": "источник",
        "сумма_дохода": "сумма",
        "тип_карты": "счёт",
        "карта": "счёт",
    }
    resolved = {}
    for key, canon in aliases.items():
        if key in mapping:
            resolved[canon] = mapping[key]
    return resolved


def is_income_sheet_transfer(source: str) -> bool:
    """Переводы между своими картами — уже на листе «Операции»."""
    s = (source or "").strip().lower()
    markers = (
        "с основной карт",
        "с общей карт",
        "с карты на кредит",
        'с карты "на кредит',
        "на вторую карт",
        "на общую карт",
        "сбер карта на кредит",
    )
    return any(m in s for m in markers)


def income_category_id(source: str, rules) -> str:
    nl = (source or "").lower()
    if "стануш" in nl:
        return "stanusha"
    if any(x in nl for x in ("контур", "дикси", "зарплат", "аванс", "премия", "отпуск", "больнич")):
        return "salary"
    if "бонус" in nl or "премия" in nl:
        return "bonus"
    if "возврат" in nl:
        return "refund"
    return suggest_category(source, rules) or "other_income"


def ensure_month(conn, ym: str, month_ids: Dict[str, str]) -> str:
    if ym in month_ids:
        return month_ids[ym]
    sort_m = conn.execute(
        "SELECT COALESCE(MAX(sort_order), 0) + 1 FROM budget_months"
    ).fetchone()[0]
    month_id = new_id()
    month_ids[ym] = month_id
    conn.execute(
        """INSERT INTO budget_months (id, year_month, sort_order, opening_balance, imported_balance, collapsed)
           VALUES (?, ?, ?, NULL, NULL, 1)""",
        (month_id, ym, sort_m),
    )
    return month_id


def import_income_sheet(ws, conn, rules, month_ids: Dict[str, str]) -> int:
    cols = read_income_headers(ws)
    for req in ("месяц", "источник", "сумма", "счёт"):
        if req not in cols:
            raise RuntimeError(f"На листе «Доходы» нет колонки «{req}» (проверьте заголовки)")

    count = 0
    skipped_zero = 0
    skipped_transfer = 0
    skipped_other = 0
    income_day = {}

    for r in range(2, ws.max_row + 1):
        ym = parse_month_cell(ws.cell(r, cols["месяц"]).value)
        source = ws.cell(r, cols["источник"]).value
        amount = parse_money(ws.cell(r, cols["сумма"]).value)
        account_raw = ws.cell(r, cols["счёт"]).value

        if not ym and not source and amount is None:
            continue
        if not ym:
            skipped_other += 1
            continue
        if amount is None or abs(amount) < 0.001:
            skipped_zero += 1
            continue

        source_str = str(source or "").strip() or "Доход"
        if is_income_sheet_transfer(source_str):
            skipped_transfer += 1
            continue

        account_id = map_account(account_raw)
        if not account_id or account_id == "credit_card":
            print(f"  доходы row {r}: неизвестная карта «{account_raw}»", file=sys.stderr)
            skipped_other += 1
            continue

        month_id = ensure_month(conn, ym, month_ids)
        day = income_day.get(ym, 0) + 1
        income_day[ym] = day
        tx_date = f"{ym}-{min(day, 28):02d}"

        max_sort = conn.execute(
            "SELECT COALESCE(MAX(sort_order), 0) FROM transactions WHERE month_id = ?",
            (month_id,),
        ).fetchone()[0]

        insert_tx(
            conn,
            id=new_id(),
            month_id=month_id,
            sort_order=max_sort + 1,
            tx_date=tx_date,
            account_id=account_id,
            target_account_id=None,
            operation_kind="regular",
            category_id=income_category_id(source_str, rules),
            income_source=source_str,
            income_amount=abs(amount),
            note="лист Доходы",
        )
        count += 1

    print(f"  доходы: {count} (пропуск: нули {skipped_zero}, переводы {skipped_transfer}, прочее {skipped_other})")
    return count


def parse_ru_date(value, last_date: Optional[str]) -> Optional[str]:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if value is None or str(value).strip() == "":
        return last_date
    s = str(value).strip().lower().replace(".", "")
    m_day = re.match(r"^(\d{1,2})\s+([а-яё]+)", s)
    if not m_day:
        return last_date
    day = int(m_day.group(1))
    month_word = m_day.group(2)
    month = None
    for part, num in MONTH_PARTS:
        if month_word.startswith(part):
            month = num
            break
    if not month:
        return last_date
    return f"{IMPORT_YEAR}-{month:02d}-{day:02d}"


def infer_operation(
    name: str,
    account_id: str,
    target_id: Optional[str],
    amount: float,
    explicit_type: Optional[str],
    tx_date: Optional[str] = None,
    credit_tracking_start: Optional[str] = None,
) -> Tuple[str, Optional[str], str]:
    """Returns (kind, target_account_id, flow) where flow is expense|income."""
    name_l = (name or "").strip().lower()
    explicit = (explicit_type or "").strip().lower()

    def credit_payment_counts() -> bool:
        if not credit_tracking_start or not tx_date:
            return True
        return tx_date >= credit_tracking_start

    if explicit in ("расход", "expense"):
        kind = "debt_payment" if account_id == "credit_payments_card" and _is_debt_name(name_l) else "regular"
        return kind, None, "expense"
    if explicit in ("доход", "income"):
        return "regular", None, "income"
    if explicit in ("перевод", "transfer"):
        return "transfer", target_id, "transfer"
    if explicit in ("корректировка", "correction"):
        return "correction", None, "correction"

    if target_id:
        if target_id == "credit_card" or target_id == account_id:
            if target_id == account_id:
                return "regular", None, "expense"
            if not credit_payment_counts():
                return "regular", None, "expense"
            return "credit_card_payment", target_id, "transfer"
        return "transfer", target_id, "transfer"

    if "на общую" in name_l:
        return "transfer", "shared_card", "transfer"
    if "на основн" in name_l and account_id == "credit_payments_card":
        return "transfer", "main_card", "transfer"
    if "на карту для оплаты кредит" in name_l or name_l in ("на кредит", "на кредиты"):
        return "transfer", "credit_payments_card", "transfer"
    if "на кредитк" in name_l:
        if not credit_payment_counts():
            return "regular", None, "expense"
        return "credit_card_payment", "credit_card", "transfer"

    if _is_income_name(name_l, amount):
        return "regular", None, "income"

    if account_id == "credit_payments_card" and _is_debt_name(name_l):
        return "debt_payment", None, "expense"

    return "regular", None, "expense"


def _is_debt_name(name_l: str) -> bool:
    keys = ("кредит", "ипотек", "автокредит", "авто сбер", "ренесанс", "совок", "страхован")
    if "стануш" in name_l:
        return False
    if "на основн" in name_l or "на общую" in name_l:
        return False
    return any(k in name_l for k in keys)


def _is_income_name(name_l: str, amount: float) -> bool:
    if any(k in name_l for k in ("стануш", "контур", "дикси", "зарплат", "аванс", "премия")):
        return True
    return False


def resolve_category(
    name: str,
    account_id: str,
    flow: str,
    slug: Optional[str],
    rules,
) -> Optional[str]:
    if slug:
        s = slug.strip().lower()
        if s in CATEGORY_SLUGS:
            return CATEGORY_SLUGS[s]
        if re.match(r"^[a-z_]+$", s):
            return s
    if flow == "income":
        nl = name.lower()
        if "стануш" in nl:
            return "stanusha"
        if any(x in nl for x in ("контур", "дикси", "зарплат", "аванс", "совок")):
            return "salary"
        return suggest_category(name, rules) or "other_income"
    if account_id == "credit_payments_card":
        return suggest_category(name, rules) or "credits"
    if account_id == "credit_card":
        return suggest_category(name, rules) or "other"
    return suggest_category(name, rules) or "other"


def read_headers(ws) -> Dict[str, int]:
    mapping = {}
    for c in range(1, ws.max_column + 1):
        h = normalize_header(ws.cell(1, c).value)
        if h:
            mapping[h] = c
    aliases = {
        "дата": "дата",
        "тип": "тип",
        "название": "название",
        "счёт": "счёт",
        "счет": "счёт",
        "счёт_куда": "счёт_куда",
        "счет_куда": "счёт_куда",
        "сумма": "сумма",
        "категория": "категория",
        "комментарий": "комментарий",
    }
    resolved = {}
    for key, canon in aliases.items():
        if key in mapping:
            resolved[canon] = mapping[key]
    return resolved


def init_db(conn):
    conn.executescript(SCHEMA_PATH.read_text(encoding="utf-8"))
    for col, defn in [("credit_limit", "REAL"), ("status", "TEXT NOT NULL DEFAULT 'active'")]:
        cols = [row[1] for row in conn.execute("PRAGMA table_info(accounts)").fetchall()]
        if col not in cols:
            conn.execute(f"ALTER TABLE accounts ADD COLUMN {col} {defn}")
    tx_cols = {
        "tx_date": "TEXT",
        "account_id": "TEXT",
        "target_account_id": "TEXT",
        "category_id": "TEXT",
        "operation_kind": "TEXT NOT NULL DEFAULT 'regular'",
        "payment_status": "TEXT NOT NULL DEFAULT 'done'",
    }
    tcols = [row[1] for row in conn.execute("PRAGMA table_info(transactions)").fetchall()]
    for col, defn in tx_cols.items():
        if col not in tcols:
            conn.execute(f"ALTER TABLE transactions ADD COLUMN {col} {defn}")
    conn.commit()


def wipe_all(conn):
    conn.execute("DELETE FROM month_category_totals")
    conn.execute("DELETE FROM transactions")
    conn.execute("DELETE FROM budget_months")
    conn.execute("DELETE FROM accounts")
    conn.execute("DELETE FROM categories")
    conn.execute("UPDATE app_settings SET import_completed_at = NULL WHERE id = 1")
    conn.commit()


def seed_accounts_categories(conn):
    for a in json.loads(SEED_ACCOUNTS.read_text(encoding="utf-8")):
        ib = float(a.get("initial_balance", 0))
        limit = a.get("credit_limit")
        if a.get("type") == "credit" and limit is not None:
            ib = float(limit)
        conn.execute(
            """INSERT INTO accounts
               (id, name, type, color, icon, initial_balance, credit_limit, status, sort_order, is_active)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1)""",
            (
                a["id"], a["name"], a["type"], a.get("color"), a.get("icon"),
                ib, limit, a.get("status", "active"), int(a.get("sort_order", 0)),
            ),
        )
    for c in json.loads(SEED_CATEGORIES.read_text(encoding="utf-8")):
        conn.execute(
            """INSERT INTO categories (id, name, type, color, icon, monthly_limit, sort_order, is_active)
               VALUES (?, ?, ?, ?, ?, ?, ?, 1)""",
            (
                c["id"], c["name"], c["type"], c.get("color"), c.get("icon"),
                c.get("monthly_limit"), int(c.get("sort_order", 0)),
            ),
        )
    conn.commit()


def read_balances_sheet(ws) -> Dict[str, dict]:
    """month ym -> snapshot"""
    col_map: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        h = normalize_header(ws.cell(1, c).value)
        if h:
            col_map[h] = c

    def col_val(row: int, *names: str):
        for name in names:
            if name in col_map:
                return parse_money(ws.cell(row, col_map[name]).value)
        return None

    result = {}
    if ws.max_row < 2:
        return result
    for r in range(2, ws.max_row + 1):
        month_raw = ws.cell(r, 1).value
        if not month_raw:
            continue
        mkey = str(month_raw).strip().lower()
        ym = None
        for part, suffix in BALANCE_MONTH_RU.items():
            if mkey.startswith(part):
                ym = f"{IMPORT_YEAR}-{suffix}"
                break
        if not ym:
            continue
        result[ym] = {
            "opening_total": col_val(r, "баланс_на_начало_месяца", "opening_total"),
            "imported_balance": col_val(
                r, "баланс_общий_на_конец_месяца", "общий_баланс", "imported_balance"
            ),
            "main_balance": col_val(r, "основная", "main_balance"),
            "shared_balance": col_val(r, "общая", "shared_balance"),
            "credit_pay_balance": col_val(r, "на_кредиты", "credit_pay_balance"),
            "credit_available": col_val(
                r, "кредитка_доступно", "кредитка", "credit_available", "доступно_кредитка"
            ),
        }
        if result[ym]["opening_total"] is None and col_map.get("месяц") != 1:
            result[ym]["opening_total"] = parse_money(ws.cell(r, 2).value)
        if result[ym]["imported_balance"] is None:
            result[ym]["imported_balance"] = parse_money(ws.cell(r, 3).value)
        if result[ym]["main_balance"] is None:
            result[ym]["main_balance"] = parse_money(ws.cell(r, 4).value)
        if result[ym]["shared_balance"] is None:
            result[ym]["shared_balance"] = parse_money(ws.cell(r, 5).value)
        if result[ym]["credit_pay_balance"] is None:
            result[ym]["credit_pay_balance"] = parse_money(ws.cell(r, 6).value)
    return result


def find_credit_tracking_start(ws, cols: Dict[str, int]) -> str:
    """С какой даты «На кредитку» уменьшает долг по кредитке (с первого месяца трат)."""
    first: Optional[str] = None
    last_date = f"{IMPORT_YEAR}-01-01"
    for r in range(2, ws.max_row + 1):
        if map_account(ws.cell(r, cols["счёт"]).value) != "credit_card":
            continue
        date_col = cols.get("дата")
        tx_date = parse_ru_date(ws.cell(r, date_col).value if date_col else None, last_date)
        if tx_date:
            last_date = tx_date
            if first is None or tx_date < first:
                first = tx_date
    if first:
        return first[:7] + "-01"
    return f"{IMPORT_YEAR}-05-01"


def insert_tx(conn, **kwargs):
    conn.execute(
        """INSERT INTO transactions
           (id, month_id, sort_order, tx_date, expense_name, expense_amount,
            income_source, income_amount, category, account_id, target_account_id,
            category_id, operation_kind, payment_status, note)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, NULL, ?, ?, ?, ?, 'done', ?)""",
        (
            kwargs["id"], kwargs["month_id"], kwargs["sort_order"], kwargs["tx_date"],
            kwargs.get("expense_name"), kwargs.get("expense_amount"),
            kwargs.get("income_source"), kwargs.get("income_amount"),
            kwargs["account_id"], kwargs.get("target_account_id"),
            kwargs.get("category_id"), kwargs["operation_kind"], kwargs.get("note", ""),
        ),
    )


def import_operations(ws, conn, rules) -> Tuple[int, Dict[str, List[dict]]]:
    cols = read_headers(ws)
    required = ("название", "счёт", "сумма")
    for req in required:
        if req not in cols:
            raise RuntimeError(f"На листе «Операции» нет колонки «{req}»")

    by_month: Dict[str, List[dict]] = {}
    last_date = f"{IMPORT_YEAR}-01-01"
    sort_global = 0
    skipped = 0
    from_fill = 0
    credit_tracking_start = find_credit_tracking_start(ws, cols)
    print(f"  учёт пополнений кредитки с {credit_tracking_start}")

    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, cols["название"]).value
        account_raw = ws.cell(r, cols["счёт"]).value
        amount = parse_money(ws.cell(r, cols["сумма"]).value)
        if not name and not account_raw and amount is None:
            continue
        if amount is None or abs(amount) < 0.001:
            skipped += 1
            continue

        account_id = map_account(account_raw)
        if not account_id:
            print(f"  skip row {r}: неизвестный счёт «{account_raw}»", file=sys.stderr)
            skipped += 1
            continue

        target_raw = ws.cell(r, cols["счёт_куда"]).value if "счёт_куда" in cols else None
        target_id = map_account(target_raw) if target_raw else None

        date_col = cols.get("дата")
        tx_date = parse_ru_date(ws.cell(r, date_col).value if date_col else None, last_date)
        if tx_date:
            last_date = tx_date
        if not tx_date or not tx_date.startswith(f"{IMPORT_YEAR}-"):
            skipped += 1
            continue

        explicit_type = None
        if "тип" in cols:
            explicit_type = ws.cell(r, cols["тип"]).value
        elif "название" in cols and cols["название"] != cols.get("тип"):
            pass
        name_str = str(name or "").strip() or "Операция"

        kind, inferred_target, flow = infer_operation(
            name_str, account_id, target_id, abs(amount), explicit_type,
            tx_date=tx_date, credit_tracking_start=credit_tracking_start,
        )
        if inferred_target and kind in ("transfer", "credit_card_payment"):
            target_id = inferred_target

        slug = ws.cell(r, cols["категория"]).value if "категория" in cols else None
        note = str(ws.cell(r, cols["комментарий"]).value or "") if "комментарий" in cols else ""
        name_col = cols["название"]
        fill_cat = category_from_fill(ws, r, name_col) if flow == "expense" else None
        if fill_cat:
            from_fill += 1
        cat_id = fill_cat or resolve_category(name_str, account_id, flow, slug, rules)

        ym = tx_date[:7]
        sort_global += 1
        tx = {
            "id": new_id(),
            "sort_order": sort_global,
            "tx_date": tx_date,
            "account_id": account_id,
            "target_account_id": target_id,
            "operation_kind": kind,
            "category_id": cat_id,
            "note": note,
        }

        if flow == "income":
            tx["income_source"] = name_str
            tx["income_amount"] = abs(amount)
        elif flow == "correction":
            tx["operation_kind"] = "correction"
            if amount < 0:
                tx["expense_name"] = name_str
                tx["expense_amount"] = abs(amount)
            else:
                tx["income_source"] = name_str
                tx["income_amount"] = abs(amount)
        else:
            tx["expense_name"] = name_str
            tx["expense_amount"] = abs(amount)

        by_month.setdefault(ym, []).append(tx)

    month_ids: Dict[str, str] = {}
    count = 0
    for sort_m, ym in enumerate(sorted(by_month.keys()), start=1):
        month_id = new_id()
        month_ids[ym] = month_id
        conn.execute(
            """INSERT INTO budget_months (id, year_month, sort_order, opening_balance, imported_balance, collapsed)
               VALUES (?, ?, ?, NULL, NULL, 1)""",
            (month_id, ym, sort_m),
        )
        for i, tx in enumerate(by_month[ym], start=1):
            tx["month_id"] = month_id
            tx["sort_order"] = i
            insert_tx(conn, **tx)
            count += 1

    if skipped:
        print(f"  пропущено строк: {skipped}")
    print(f"  категория по заливке: {from_fill}")
    return count, month_ids


def apply_balance_snapshots(conn, snapshots: Dict[str, dict]):
    for ym, snap in snapshots.items():
        row = conn.execute(
            "SELECT id FROM budget_months WHERE year_month = ?", (ym,)
        ).fetchone()
        if not row:
            month_id = new_id()
            sort_order = int(ym.replace("-", ""))
            conn.execute(
                """INSERT INTO budget_months (id, year_month, sort_order, opening_balance, imported_balance, collapsed)
                   VALUES (?, ?, ?, ?, ?, 1)""",
                (month_id, ym, sort_order, snap.get("opening_total"), snap.get("imported_balance")),
            )
        else:
            conn.execute(
                """UPDATE budget_months
                   SET opening_balance = COALESCE(?, opening_balance),
                       imported_balance = COALESCE(?, imported_balance)
                   WHERE year_month = ?""",
                (snap.get("opening_total"), snap.get("imported_balance"), ym),
            )

    jan = snapshots.get(f"{IMPORT_YEAR}-01")
    if jan and jan.get("opening_total") is not None:
        conn.execute(
            "UPDATE app_settings SET initial_opening_balance = ? WHERE id = 1",
            (jan["opening_total"],),
        )
        conn.execute(
            "UPDATE accounts SET initial_balance = ? WHERE id = 'main_card'",
            (jan["opening_total"],),
        )
        conn.execute(
            "UPDATE budget_months SET opening_balance = ? WHERE year_month = ?",
            (jan["opening_total"], f"{IMPORT_YEAR}-01"),
        )

    # до апреля всё на основной карте
    for acc in ("shared_card", "credit_payments_card"):
        conn.execute("UPDATE accounts SET initial_balance = 0 WHERE id = ?", (acc,))


def _tx_amount(tx) -> float:
    return abs(tx["expense_amount"] or 0) or abs(tx["income_amount"] or 0)


def _account_delta(tx, account_id: str) -> float:
    if tx["payment_status"] == "ignored":
        return 0.0
    delta = 0.0
    amount = _tx_amount(tx)
    kind = tx["operation_kind"]
    if tx["account_id"] == account_id:
        if kind == "correction":
            if tx["income_amount"]:
                delta += tx["income_amount"]
            else:
                delta -= tx["expense_amount"] or 0
        elif kind in ("transfer", "credit_card_payment"):
            delta -= amount
        elif (tx["expense_amount"] or 0) > 0:
            delta -= tx["expense_amount"]
        elif (tx["income_amount"] or 0) > 0:
            delta += tx["income_amount"]
    if tx["target_account_id"] == account_id and kind in ("transfer", "credit_card_payment"):
        delta += amount
    return delta


def _closing_balances(conn, month_id: str) -> Dict[str, float]:
    accounts = {
        r[0]: {"type": r[1], "initial": r[2], "limit": r[3]}
        for r in conn.execute("SELECT id, type, initial_balance, credit_limit FROM accounts")
    }
    months = [r[0] for r in conn.execute(
        "SELECT id FROM budget_months ORDER BY sort_order"
    )]
    running = {aid: data["initial"] for aid, data in accounts.items()}
    result = {aid: running[aid] for aid in running}

    for mid in months:
        txs = conn.execute(
            """SELECT account_id, target_account_id, operation_kind, payment_status,
                      expense_amount, income_amount, category_id
               FROM transactions WHERE month_id = ? ORDER BY sort_order""",
            (mid,),
        ).fetchall()
        for tx in txs:
            txd = {
                "account_id": tx[0], "target_account_id": tx[1], "operation_kind": tx[2],
                "payment_status": tx[3], "expense_amount": tx[4], "income_amount": tx[5],
                "category_id": tx[6],
            }
            for aid, data in accounts.items():
                if data["type"] == "credit":
                    if txd["account_id"] == aid and txd["operation_kind"] == "correction":
                        if txd["income_amount"]:
                            running[aid] += txd["income_amount"]
                        else:
                            running[aid] -= txd["expense_amount"] or 0
                    elif (
                        txd["target_account_id"] == aid
                        and txd["operation_kind"] in ("transfer", "credit_card_payment")
                    ):
                        running[aid] += _tx_amount(txd)
                    elif (
                        txd["account_id"] == aid
                        and txd["operation_kind"] == "regular"
                        and (txd["expense_amount"] or 0) > 0
                        and txd.get("category_id") in ("transfer", "credit_card_payment")
                    ):
                        running[aid] += txd["expense_amount"]
                    elif (
                        txd["account_id"] == aid
                        and txd["operation_kind"] == "regular"
                        and (txd["income_amount"] or 0) > 0
                        and not (txd["expense_amount"] or 0)
                    ):
                        running[aid] += txd["income_amount"]
                    elif txd["account_id"] == aid and txd["operation_kind"] == "regular" and (txd["expense_amount"] or 0) > 0:
                        running[aid] -= txd["expense_amount"]
                else:
                    running[aid] += _account_delta(txd, aid)
        result = {aid: running[aid] for aid in running}
        if mid == month_id:
            break

    out = {}
    for aid, data in accounts.items():
        if data["type"] == "credit":
            limit = data["limit"] or 0
            out[aid] = running[aid] - limit
        else:
            out[aid] = running[aid]
    return out


def reconcile_with_snapshots(conn, snapshots: Dict[str, dict]) -> int:
    """Корректировки на конец месяца по листу «Балансы»."""
    added = 0
    debit_targets = {
        "main_card": "main_balance",
        "shared_card": "shared_balance",
        "credit_payments_card": "credit_pay_balance",
    }
    credit_reconcile_ym = None
    for ym in sorted(snapshots.keys(), reverse=True):
        if snapshots[ym].get("credit_available") is not None:
            credit_reconcile_ym = ym
            break

    limit_row = conn.execute(
        "SELECT credit_limit FROM accounts WHERE id = 'credit_card'"
    ).fetchone()
    credit_limit = float(limit_row[0] or 0) if limit_row else 380000.0

    for ym in sorted(snapshots.keys()):
        snap = snapshots[ym]
        row = conn.execute("SELECT id FROM budget_months WHERE year_month = ?", (ym,)).fetchone()
        if not row:
            continue
        month_id = row[0]
        last_day = f"{ym}-28"
        if ym.endswith("-01") or ym.endswith("-03") or ym.endswith("-05") or ym.endswith("-07") or ym.endswith("-08") or ym.endswith("-10") or ym.endswith("-12"):
            last_day = f"{ym}-31"
        elif ym.endswith("-02"):
            last_day = f"{ym}-28"
        else:
            last_day = f"{ym}-30"

        computed = _closing_balances(conn, month_id)
        max_sort = conn.execute(
            "SELECT COALESCE(MAX(sort_order), 0) FROM transactions WHERE month_id = ?",
            (month_id,),
        ).fetchone()[0]

        for acc_id, snap_key in debit_targets.items():
            expected = snap.get(snap_key)
            if expected is None:
                continue
            diff = expected - computed.get(acc_id, 0)
            if abs(diff) < 1:
                continue
            max_sort += 1
            tx = {
                "id": new_id(),
                "month_id": month_id,
                "sort_order": max_sort,
                "tx_date": last_day,
                "account_id": acc_id,
                "target_account_id": None,
                "operation_kind": "correction",
                "category_id": "correction",
                "note": "сверка с листом Балансы",
            }
            if diff > 0:
                tx["income_source"] = "Сверка баланса"
                tx["income_amount"] = diff
            else:
                tx["expense_name"] = "Сверка баланса"
                tx["expense_amount"] = abs(diff)
            insert_tx(conn, **tx)
            added += 1

        if ym == credit_reconcile_ym:
            expected_avail = snap.get("credit_available")
            if expected_avail is not None:
                expected_balance = expected_avail - credit_limit
                diff = expected_balance - computed.get("credit_card", 0)
                if abs(diff) >= 1:
                    max_sort += 1
                    tx = {
                        "id": new_id(),
                        "month_id": month_id,
                        "sort_order": max_sort,
                        "tx_date": last_day,
                        "account_id": "credit_card",
                        "target_account_id": None,
                        "operation_kind": "correction",
                        "category_id": "correction",
                        "note": "сверка доступного лимита кредитки",
                    }
                    if diff > 0:
                        tx["income_source"] = "Сверка кредитки"
                        tx["income_amount"] = diff
                    else:
                        tx["expense_name"] = "Сверка кредитки"
                        tx["expense_amount"] = abs(diff)
                    insert_tx(conn, **tx)
                    added += 1

        total_expected = snap.get("imported_balance")
        if total_expected is None:
            total_expected = sum(
                snap.get(k) or 0 for k in ("main_balance", "shared_balance", "credit_pay_balance")
            )
        conn.execute(
            "UPDATE budget_months SET imported_balance = ? WHERE id = ?",
            (total_expected, month_id),
        )

    return added


def apply_credit_available_override(snapshots: Dict[str, dict], credit_limit: float) -> None:
    ym = None
    for candidate in sorted(snapshots.keys(), reverse=True):
        if snapshots[candidate].get("credit_available") is not None:
            ym = candidate
            break
    if not ym:
        return
    avail = snapshots[ym]["credit_available"]
    if avail is None or abs(avail - credit_limit) < 1:
        snapshots[ym]["credit_available"] = CREDIT_AVAILABLE_FALLBACK


def import_xlsx(path: Path):
    rules = load_rules()
    conn = sqlite3.connect(DB_PATH)
    init_db(conn)
    wipe_all(conn)
    seed_accounts_categories(conn)

    wb = openpyxl.load_workbook(path, data_only=False)
    if "Операции" not in wb.sheetnames:
        raise RuntimeError("Нет листа «Операции»")

    tx_count, month_ids = import_operations(wb["Операции"], conn, rules)

    income_count = 0
    if "Доходы" in wb.sheetnames:
        income_count = import_income_sheet(wb["Доходы"], conn, rules, month_ids)

    snapshots = {}
    if "Балансы" in wb.sheetnames:
        snapshots = read_balances_sheet(wb["Балансы"])
        apply_credit_available_override(snapshots, CREDIT_LIMIT_DEFAULT)
        apply_balance_snapshots(conn, snapshots)
        corrections = reconcile_with_snapshots(conn, snapshots)
        if corrections:
            print(f"  корректировок по балансам: {corrections}")

    conn.execute(
        "UPDATE app_settings SET import_completed_at = ? WHERE id = 1",
        (datetime.utcnow().isoformat() + "Z",),
    )
    conn.commit()

    months = conn.execute("SELECT COUNT(*) FROM budget_months").fetchone()[0]
    total_tx = conn.execute("SELECT COUNT(*) FROM transactions").fetchone()[0]
    print(f"Импорт: месяцев {months}, операций {total_tx} (расходы {tx_count}, доходы {income_count})")

    if snapshots:
        print("Сверка балансов (дебетовые карты, без кредитки):")
        for ym in sorted(snapshots.keys()):
            row = conn.execute("SELECT id FROM budget_months WHERE year_month = ?", (ym,)).fetchone()
            if not row:
                continue
            closing = _closing_balances(conn, row[0])
            total = sum(closing[aid] for aid in ("main_card", "shared_card", "credit_payments_card"))
            snap = snapshots[ym]
            expected = snap.get("imported_balance")
            if expected is None:
                expected = sum(snap.get(k) or 0 for k in ("main_balance", "shared_balance", "credit_pay_balance"))
            ok = "OK" if abs(total - expected) < 2 else f"Δ{total - expected:.0f}"
            print(f"  {ym}: {total:.0f} / {expected:.0f} {ok}")
        june_row = conn.execute(
            "SELECT id FROM budget_months WHERE year_month = ?", (f"{IMPORT_YEAR}-06",)
        ).fetchone()
        if june_row:
            closing = _closing_balances(conn, june_row[0])
            cc = closing.get("credit_card", 0)
            print(f"  кредитка: баланс {cc:.0f} (доступно {cc + 380000:.0f}, лимит 380000)")

    conn.close()


if __name__ == "__main__":
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    import_xlsx(xlsx)
