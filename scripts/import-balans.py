#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Импорт Баланс.xlsx — только год IMPORT_YEAR (по умолчанию 2026).

C — расход основная    D — расход общая    E — расход на кредиты    F — расход кредитка
G — источник (основная) H — пополнение основной
K — пополнение на кредиты   L — пополнение общая

Итого (строка сразу после заголовка месяца):
  B — общие расходы   C/D/E/F — расходы по картам   H — доходы основной
  I — общий баланс (искл. 2026-01 → J)   J — баланс основной   K — общая   L — на кредиты
"""

import json
import sqlite3
import sys
import uuid
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent
DB_PATH = ROOT / "data" / "personal-budget.sqlite"
SCHEMA_PATH = ROOT / "api" / "schema.sql"
RULES_PATH = ROOT / "api" / "rules" / "categorize.json"
SEED_ACCOUNTS = ROOT / "api" / "seed-accounts.json"
SEED_CATEGORIES = ROOT / "api" / "seed-categories.json"

# Расходы по картам (столбцы C–F)
EXPENSE_BY_COL = {
    3: "main_card",
    4: "shared_card",
    5: "credit_payments_card",
    6: "credit_card",
}

# Пополнения: (столбец суммы, столбец источника или None)
INCOME_BY_COL = {
    8: ("main_card", 7),
    11: ("credit_payments_card", None),
    12: ("shared_card", None),
}

EXCEL_CATEGORIES = [
    (13, "Еда"),
    (14, "Ежемесячные"),
    (15, "Стануша"),
    (16, "Кредиты"),
    (17, "Остальное"),
    (18, "Кредитка"),
    (19, "Алкоголь"),
]

IMPORT_YEAR = 2026

MONTH_PARTS = [
    ("январ", 1), ("феврал", 2), ("март", 3), ("апрел", 4),
    ("май", 5), ("июн", 6), ("июл", 7), ("август", 8),
    ("сентябр", 9), ("октябр", 10), ("ноябр", 11), ("декабр", 12),
]


def parse_money(value) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str) and value.startswith("#"):
        return None
    s = str(value).replace("\xa0", "").replace(" ", "").replace("₽", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def cell_string(value) -> Optional[str]:
    if value is None or value == "":
        return None
    return str(value).strip()


def new_id() -> str:
    return str(uuid.uuid4())


def load_rules():
    if RULES_PATH.is_file():
        return json.loads(RULES_PATH.read_text(encoding="utf-8"))
    return []


def suggest_category(title: str, rules) -> Optional[str]:
    lower = title.lower()
    for rule in rules:
        for kw in rule.get("keywords", []):
            if kw.lower() in lower:
                return rule["categoryId"]
    return None


def parse_month_header(value) -> Optional[str]:
    """Только текстовые заголовки «Июнь 2024», не даты в операциях."""
    if not isinstance(value, str):
        return None
    s = value.strip().lower()
    if s.startswith("итого") or "доход" in s or "расход" in s:
        return None
    year = None
    for part in s.replace(".", " ").split():
        if part.isdigit() and len(part) == 4:
            year = int(part)
    if not year:
        return None
    for name, month in MONTH_PARTS:
        if name in s:
            return f"{year}-{month:02d}"
    return None


def is_itogo_row(ws, r: int) -> bool:
    a = ws.cell(r, 1).value
    return bool(a and str(a).strip().lower().startswith("итого"))


def find_month_sections(ws) -> List[dict]:
    """Заголовок месяца → строка Итого → операции до следующего заголовка."""
    seen: Dict[str, int] = {}
    sections = []
    for r in range(1, ws.max_row + 1):
        ym = parse_month_header(ws.cell(r, 1).value)
        if not ym or ym in seen:
            continue
        if not is_itogo_row(ws, r + 1):
            continue
        seen[ym] = r
        sections.append({"year_month": ym, "header_row": r, "itogo_row": r + 1})

    for i, sec in enumerate(sections):
        sec["tx_start"] = sec["itogo_row"] + 1
        sec["tx_end"] = (
            sections[i + 1]["header_row"] - 1 if i + 1 < len(sections) else ws.max_row
        )
    return sections


def read_itogo(ws, itogo_row: int, year_month: str) -> dict:
    i_bal = parse_money(ws.cell(itogo_row, 9).value)
    j_bal = parse_money(ws.cell(itogo_row, 10).value)
    # Исключение пользователя: общий баланс января 2026 в J
    if year_month == "2026-01" and j_bal is not None:
        total_balance = j_bal
    else:
        total_balance = i_bal

    return {
        "total_expenses": parse_money(ws.cell(itogo_row, 2).value),
        "main_expense": parse_money(ws.cell(itogo_row, 3).value),
        "shared_expense": parse_money(ws.cell(itogo_row, 4).value),
        "credit_pay_expense": parse_money(ws.cell(itogo_row, 5).value),
        "credit_card_expense": parse_money(ws.cell(itogo_row, 6).value),
        "main_income": parse_money(ws.cell(itogo_row, 8).value),
        "total_balance": total_balance,
        "main_balance": j_bal,
        "shared_balance": parse_money(ws.cell(itogo_row, 11).value),
        "credit_pay_balance": parse_money(ws.cell(itogo_row, 12).value),
    }


def init_db(conn):
    conn.executescript(SCHEMA_PATH.read_text(encoding="utf-8"))
    for col, defn in [("credit_limit", "REAL"), ("status", "TEXT NOT NULL DEFAULT 'active'")]:
        cols = [row[1] for row in conn.execute("PRAGMA table_info(accounts)").fetchall()]
        if col not in cols:
            conn.execute(f"ALTER TABLE accounts ADD COLUMN {col} {defn}")
    tx_cols = {
        "tx_date": "TEXT",
        "account_id": "TEXT",
        "target_account_id": "TEXT",
        "category_id": "TEXT",
        "operation_kind": "TEXT NOT NULL DEFAULT 'regular'",
        "payment_status": "TEXT NOT NULL DEFAULT 'done'",
    }
    tcols = [row[1] for row in conn.execute("PRAGMA table_info(transactions)").fetchall()]
    for col, defn in tx_cols.items():
        if col not in tcols:
            conn.execute(f"ALTER TABLE transactions ADD COLUMN {col} {defn}")
    conn.commit()


def wipe_all(conn):
    conn.execute("DELETE FROM month_category_totals")
    conn.execute("DELETE FROM transactions")
    conn.execute("DELETE FROM budget_months")
    conn.execute("DELETE FROM accounts")
    conn.execute("DELETE FROM categories")
    conn.execute("UPDATE app_settings SET import_completed_at = NULL WHERE id = 1")
    conn.commit()


def seed_accounts_categories(conn):
    for a in json.loads(SEED_ACCOUNTS.read_text(encoding="utf-8")):
        ib = float(a.get("initial_balance", 0))
        limit = a.get("credit_limit")
        if a.get("type") == "credit" and limit is not None:
            ib = float(limit)
        conn.execute(
            """INSERT INTO accounts
               (id, name, type, color, icon, initial_balance, credit_limit, status, sort_order, is_active)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1)""",
            (
                a["id"], a["name"], a["type"], a.get("color"), a.get("icon"),
                ib, limit, a.get("status", "active"), int(a.get("sort_order", 0)),
            ),
        )
    for c in json.loads(SEED_CATEGORIES.read_text(encoding="utf-8")):
        conn.execute(
            """INSERT INTO categories (id, name, type, color, icon, monthly_limit, sort_order, is_active)
               VALUES (?, ?, ?, ?, ?, ?, ?, 1)""",
            (
                c["id"], c["name"], c["type"], c.get("color"), c.get("icon"),
                c.get("monthly_limit"), int(c.get("sort_order", 0)),
            ),
        )
    conn.commit()


def insert_tx(
    conn,
    month_id: str,
    sort_order: int,
    tx_date: Optional[str],
    account_id: str,
    kind: str,
    category_id: Optional[str],
    expense_name=None,
    expense_amount=None,
    income_source=None,
    income_amount=None,
):
    conn.execute(
        """INSERT INTO transactions
           (id, month_id, sort_order, tx_date, expense_name, expense_amount,
            income_source, income_amount, category, account_id, target_account_id,
            category_id, operation_kind, payment_status, note)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, NULL, ?, NULL, ?, ?, 'done', '')""",
        (
            new_id(), month_id, sort_order, tx_date, expense_name, expense_amount,
            income_source, income_amount, account_id, category_id, kind,
        ),
    )


def income_category(source: str, rules) -> str:
    s = (source or "").lower()
    if "стануша" in s:
        return "stanusha"
    if any(x in s for x in ("контур", "дикси", "зарплат", "аванс")):
        return "salary"
    return suggest_category(source, rules) or "other_income"


def expense_category(name: str, account_id: str, rules) -> str:
    if account_id == "credit_payments_card":
        return suggest_category(name, rules) or "credits"
    if account_id == "credit_card":
        return suggest_category(name, rules) or "other"
    return suggest_category(name, rules) or "other"


def import_transactions(ws, conn, sections: List[dict], rules) -> int:
    count = 0
    month_ids: Dict[str, str] = {}

    for sort_month, sec in enumerate(sections, start=1):
        ym = sec["year_month"]
        itogo = read_itogo(ws, sec["itogo_row"], ym)
        month_id = new_id()
        month_ids[ym] = month_id

        opening = itogo["total_balance"] if sort_month == 1 else None
        conn.execute(
            """INSERT INTO budget_months (id, year_month, sort_order, opening_balance, imported_balance, collapsed)
               VALUES (?, ?, ?, ?, ?, 1)""",
            (month_id, ym, sort_month, opening, itogo["total_balance"]),
        )

        for col, cat_name in EXCEL_CATEGORIES:
            amount = parse_money(ws.cell(sec["itogo_row"], col).value)
            if amount is not None and abs(amount) >= 0.001:
                conn.execute(
                    "INSERT INTO month_category_totals (month_id, category, amount) VALUES (?, ?, ?)",
                    (month_id, cat_name, amount),
                )

        tx_order = 0
        last_date = f"{ym}-01"

        for r in range(sec["tx_start"], sec["tx_end"] + 1):
            a = ws.cell(r, 1).value
            if parse_month_header(a) or is_itogo_row(ws, r):
                break

            if isinstance(a, datetime):
                last_date = a.strftime("%Y-%m-%d")
                tx_date = last_date
            else:
                tx_date = last_date

            # Только операции с датой в IMPORT_YEAR
            if tx_date and not tx_date.startswith(f"{IMPORT_YEAR}-"):
                continue

            name_b = cell_string(ws.cell(r, 2).value)

            # Расходы C/D/E/F
            for col, account_id in EXPENSE_BY_COL.items():
                amt = parse_money(ws.cell(r, col).value)
                if amt is None or abs(amt) < 0.001:
                    continue
                title = name_b or "Расход"
                cat = expense_category(title, account_id, rules)
                kind = "debt_payment" if account_id == "credit_payments_card" else "regular"
                tx_order += 1
                insert_tx(
                    conn, month_id, tx_order, tx_date, account_id, kind, cat,
                    expense_name=title, expense_amount=amt,
                )
                count += 1

            # Пополнения H / K / L
            for col, (account_id, src_col) in INCOME_BY_COL.items():
                amt = parse_money(ws.cell(r, col).value)
                if amt is None or abs(amt) < 0.001:
                    continue
                source = cell_string(ws.cell(r, src_col).value) if src_col else (name_b or "Пополнение")
                if not source:
                    source = "Пополнение"
                cat = income_category(source, rules)
                tx_order += 1
                insert_tx(
                    conn, month_id, tx_order, tx_date, account_id, "regular", cat,
                    income_source=source, income_amount=amt,
                )
                count += 1

    return count


def apply_opening_balances(ws, all_sections: List[dict], conn) -> None:
    """Стартовые остатки карт — из итого декабря года до IMPORT_YEAR."""
    prior_ym = f"{IMPORT_YEAR - 1}-12"
    prior = next((s for s in all_sections if s["year_month"] == prior_ym), None)
    if prior:
        snap = read_itogo(ws, prior["itogo_row"], prior_ym)
        if snap["main_balance"] is not None:
            conn.execute(
                "UPDATE accounts SET initial_balance = ? WHERE id = 'main_card'",
                (snap["main_balance"],),
            )
        if snap["shared_balance"] is not None:
            conn.execute(
                "UPDATE accounts SET initial_balance = ? WHERE id = 'shared_card'",
                (snap["shared_balance"],),
            )
        if snap["credit_pay_balance"] is not None:
            conn.execute(
                "UPDATE accounts SET initial_balance = ? WHERE id = 'credit_payments_card'",
                (snap["credit_pay_balance"],),
            )
        total = snap["total_balance"]
        if total is not None:
            conn.execute(
                "UPDATE app_settings SET initial_opening_balance = ? WHERE id = 1",
                (total,),
            )
    else:
        first = next((s for s in all_sections if s["year_month"] == f"{IMPORT_YEAR}-01"), None)
        if first:
            it0 = read_itogo(ws, first["itogo_row"], first["year_month"])
            if it0["main_balance"] is not None:
                conn.execute(
                    "UPDATE accounts SET initial_balance = ? WHERE id = 'main_card'",
                    (it0["main_balance"],),
                )
            conn.execute(
                "UPDATE app_settings SET initial_opening_balance = ? WHERE id = 1",
                (it0["total_balance"] or 0,),
            )


def import_xlsx(path: str):
    path = Path(path)
    if not path.is_file():
        raise FileNotFoundError(path)

    rules = load_rules()
    conn = sqlite3.connect(DB_PATH)
    init_db(conn)
    wipe_all(conn)
    seed_accounts_categories(conn)

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Баланс"]
    all_sections = find_month_sections(ws)
    sections = [s for s in all_sections if s["year_month"].startswith(f"{IMPORT_YEAR}-")]
    if not sections:
        raise RuntimeError(f"Не найдены месяцы за {IMPORT_YEAR}")

    tx_count = import_transactions(ws, conn, sections, rules)
    apply_opening_balances(ws, all_sections, conn)

    conn.execute(
        "UPDATE app_settings SET import_completed_at = ? WHERE id = 1",
        (datetime.utcnow().isoformat() + "Z",),
    )
    conn.commit()

    total_tx = conn.execute("SELECT COUNT(*) FROM transactions").fetchone()[0]
    june = conn.execute(
        "SELECT imported_balance FROM budget_months WHERE year_month = '2026-06'"
    ).fetchone()
    conn.close()

    print(f"Год {IMPORT_YEAR}: месяцев {len(sections)}, операций {total_tx} (вставлено {tx_count})")
    if june:
        print(f"Общий баланс 2026-06 (I3230): {june[0]}")
    # контрольные точки из ТЗ
    checks = [
        ("2026-06", "I", 3230, 21093),
        ("2026-05", "I", 3097, 19856),
        ("2026-04", "I", 2948, 58763),
    ]
    wb2 = openpyxl.load_workbook(path, data_only=True)
    ws2 = wb2["Баланс"]
    for ym, col_letter, row, expected in checks:
        col = openpyxl.utils.column_index_from_string(col_letter)
        actual = parse_money(ws2.cell(row, col).value)
        ok = "OK" if actual == expected else f"ожидалось {expected}"
        print(f"  {ym} {col_letter}{row}: {actual} {ok}")


if __name__ == "__main__":
    xlsx = sys.argv[1] if len(sys.argv) > 1 else r"c:\Users\Вячеслав\Desktop\Баланс.xlsx"
    import_xlsx(xlsx)
