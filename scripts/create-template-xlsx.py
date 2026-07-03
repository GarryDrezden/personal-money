#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Создаёт пустые Excel-шаблоны для ручного импорта."""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / "data" / "template"

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill("solid", fgColor="E8F0FE")
WRAP = Alignment(wrap_text=True, vertical="top")


def style_header_row(ws, widths):
    for col, width in enumerate(widths, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28


def build_operations_sheet(ws):
    headers = [
        "дата",
        "тип",
        "счёт",
        "счёт_куда",
        "сумма",
        "название",
        "категория",
        "комментарий",
    ]
    ws.append(headers)
    style_header_row(ws, [12, 14, 14, 14, 10, 28, 16, 24])
    ws.cell(row=2, column=1).number_format = "yyyy-mm-dd"


def build_balances_sheet(ws):
    headers = [
        "месяц",
        "общий_баланс",
        "основная",
        "общая",
        "на_кредиты",
        "кредитка_доступно",
    ]
    ws.append(headers)
    style_header_row(ws, [12, 14, 12, 12, 14, 18])


def build_help_sheet(ws):
    ws.column_dimensions["A"].width = 100
    lines = [
        "ШАБЛОН ДЛЯ PERSONAL BUDGET — заполняйте лист «Операции» (обязательно).",
        "",
        "Лист «Операции» — одна строка = одна операция, только 2026 год.",
        "",
        "дата — 2026-06-15",
        "тип — расход | доход | перевод | корректировка",
        "счёт — основная | общая | на_кредиты | кредитка",
        "счёт_куда — только для перевода (куда ушли деньги)",
        "сумма — положительное число; для корректировки можно минус",
        "название — Пятёрочка, Контур, На общую карту…",
        "категория — можно пусто (еда, зарплата, кредиты, игры…)",
        "",
        "Лист «Балансы» — итоги на конец месяца для сверки (необязательно).",
        "",
        "Сохраните файл как .xlsx и пришлите в чат с текстом «импортируй шаблон».",
    ]
    for i, text in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.alignment = WRAP
        if i == 1:
            cell.font = Font(bold=True, size=12)


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUT_DIR / "ШАБЛОН_personal_budget.xlsx"

    wb = openpyxl.Workbook()
    ws_ops = wb.active
    ws_ops.title = "Операции"
    build_operations_sheet(ws_ops)

    ws_bal = wb.create_sheet("Балансы")
    build_balances_sheet(ws_bal)

    ws_help = wb.create_sheet("Справка")
    build_help_sheet(ws_help)

    wb.save(path)
    print(f"Created: {path}")


if __name__ == "__main__":
    main()
