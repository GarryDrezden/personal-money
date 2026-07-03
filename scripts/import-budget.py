#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Import Бюджет.xlsx into personal-budget SQLite."""

import json
import re
import sqlite3
import sys
import uuid
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

CATEGORIES = [
    "Основная карта",
    "На кредиты",
    "Общая карта",
    "Еда",
    "Ежемесячные",
    "Стануша",
    "Кредиты",
    "Остальное",
    "Кредитка",
    "Алкоголь",
]

ROOT = Path(__file__).resolve().parent.parent
DB_PATH = ROOT / "data" / "personal-budget.sqlite"
SCHEMA_PATH = ROOT / "api" / "schema.sql"


def parse_money(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).replace("\xa0", "").replace(" ", "").replace("₽", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def cell_string(value):
    if value is None or value == "":
        return None
    return str(value).strip()


def new_id():
    return str(uuid.uuid4())


def init_db(conn):
    conn.executescript(SCHEMA_PATH.read_text(encoding="utf-8"))
    conn.commit()


def wipe(conn):
    conn.execute("DELETE FROM month_category_totals")
    conn.execute("DELETE FROM transactions")
    conn.execute("DELETE FROM budget_months")
    conn.execute("UPDATE app_settings SET import_completed_at = NULL WHERE id = 1")
    conn.commit()


def import_xlsx(path: str, force: bool = False):
    path = Path(path)
    if not path.is_file():
        raise FileNotFoundError(path)

    conn = sqlite3.connect(DB_PATH)
    init_db(conn)

    row = conn.execute("SELECT import_completed_at, initial_opening_balance FROM app_settings WHERE id = 1").fetchone()
    if row and row[0] and not force:
        raise RuntimeError("Import already completed. Use --force")

    opening = float(row[1]) if row and row[1] is not None else 2007.0

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    month_starts = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, datetime):
            month_starts.append(r)

    month_starts.append(ws.max_row + 1)

    wipe(conn)
    sort_order = 0
    tx_count = 0

    for idx in range(len(month_starts) - 1):
        start = month_starts[idx]
        end = month_starts[idx + 1]
        dt = ws.cell(start, 1).value
        year_month = dt.strftime("%Y-%m")
        sort_order += 1
        month_id = new_id()

        total_row = start + 1
        imported_balance = parse_money(ws.cell(total_row, 9).value)
        opening_balance = opening if sort_order == 1 else None

        conn.execute(
            """INSERT INTO budget_months (id, year_month, sort_order, opening_balance, imported_balance, collapsed)
               VALUES (?, ?, ?, ?, ?, 1)""",
            (month_id, year_month, sort_order, opening_balance, imported_balance),
        )

        for i, cat in enumerate(CATEGORIES):
            col = 10 + i
            amount = parse_money(ws.cell(total_row, col).value)
            if amount is not None and abs(amount) >= 0.001:
                conn.execute(
                    "INSERT INTO month_category_totals (month_id, category, amount) VALUES (?, ?, ?)",
                    (month_id, cat, amount),
                )

        tx_order = 0
        for r in range(start + 2, end):
            a = ws.cell(r, 1).value
            if a and str(a).startswith("Итого"):
                continue
            expense_name = cell_string(ws.cell(r, 2).value)
            expense_amount = parse_money(ws.cell(r, 3).value)
            income_source = cell_string(ws.cell(r, 7).value)
            income_amount = parse_money(ws.cell(r, 8).value)
            if not any([expense_name, expense_amount, income_source, income_amount]):
                continue
            tx_order += 1
            conn.execute(
                """INSERT INTO transactions
                   (id, month_id, sort_order, expense_name, expense_amount, income_source, income_amount, category, note)
                   VALUES (?, ?, ?, ?, ?, ?, ?, NULL, '')""",
                (new_id(), month_id, tx_order, expense_name, expense_amount, income_source, income_amount),
            )
            tx_count += 1

    conn.execute(
        "UPDATE app_settings SET import_completed_at = ? WHERE id = 1",
        (datetime.utcnow().isoformat() + "Z",),
    )
    conn.commit()
    conn.close()

    return {"months": sort_order, "transactions": tx_count}


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python scripts/import-budget.py <file.xlsx> [--force]", file=sys.stderr)
        sys.exit(1)
    force = "--force" in sys.argv
    result = import_xlsx(sys.argv[1], force)
    print(json.dumps(result, ensure_ascii=False, indent=2))
